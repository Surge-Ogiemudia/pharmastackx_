import openpyxl
import json
import sys

file_path = r"C:\Users\HP\Desktop\Sales and Inventory Management and Information System  Macrosales.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    print("Sheet names:", wb.sheetnames)
    
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    print(f"Total rows in active sheet '{sheet.title}': {len(rows)}")
    
    if len(rows) > 0:
        print("First 10 rows:")
        for idx, row in enumerate(rows[:10]):
            print(f"Row {idx}: {row}")
except Exception as e:
    print("Error loading excel:", e)
